import os
import logging
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from supabase import create_client, Client
from datetime import datetime, date, timedelta
import pytz
import requests
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from collections import defaultdict

# Импорт openpyxl с защитой
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Excel export disabled.")

logging.basicConfig(level=logging.INFO)

app = Flask(__name__, template_folder='../templates')
CORS(app)

# ========== ПЕРЕМЕННЫЕ ОКРУЖЕНИЯ ==========
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
ADMIN_CHAT_ID = os.environ.get("ADMIN_CHAT_ID")

SMTP_HOST = os.environ.get("SMTP_HOST")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 587)) if os.environ.get("SMTP_PORT") else 587
SMTP_USER = os.environ.get("SMTP_USER")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USER)

# ========== SUPABASE ==========
if not SUPABASE_URL or not SUPABASE_KEY:
    logging.error("Missing SUPABASE_URL or SUPABASE_KEY")
    supabase = None
else:
    try:
        supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
        logging.info("Supabase client initialized")
    except Exception as e:
        logging.error(f"Supabase init error: {e}")
        supabase = None

if not TELEGRAM_BOT_TOKEN:
    logging.warning("TELEGRAM_BOT_TOKEN not set")

YEKAT_TIMEZONE = pytz.timezone('Asia/Yekaterinburg')

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
def get_gosb_by_slug(slug):
    if not supabase:
        return None
    try:
        res = supabase.table('gosb').select('*').eq('slug', slug).execute()
        return res.data[0] if res.data else None
    except Exception as e:
        logging.error(f"get_gosb_by_slug error: {e}")
        return None

def get_cities_by_gosb(gosb_id):
    if not supabase:
        return []
    try:
        res = supabase.table('cities').select('id, name').eq('gosb_id', gosb_id).execute()
        return res.data
    except Exception as e:
        logging.error(f"get_cities_by_gosb error: {e}")
        return []

def reverse_geocode(lat, lng):
    try:
        lat = float(lat); lng = float(lng)
        lat_fixed = round(lat, 5); lng_fixed = round(lng, 5)
        url = f"https://nominatim.openstreetmap.org/reverse?format=jsonv2&lat={lat_fixed}&lon={lng_fixed}&accept-language=ru&zoom=18&addressdetails=1"
        headers = {'User-Agent': 'Mozilla/5.0 (compatible; RegistrationApp/1.0)'}
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code == 200:
            data = response.json()
            if data and data.get('address'):
                addr = data['address']
                parts = []
                road = addr.get('road')
                house = addr.get('house_number')
                if road: parts.append(road)
                if house: parts.append('д. ' + house)
                city = addr.get('city') or addr.get('town') or addr.get('village') or addr.get('hamlet')
                if city:
                    if parts: parts.insert(0, city + ',')
                    else: parts.append(city)
                if len(parts) < 2 and data.get('display_name'):
                    display = data['display_name'].split(', Россия')[0]
                    display_parts = display.split(', ')
                    parts = [display_parts[0] + ',', display_parts[1]] if len(display_parts) >= 2 else [display]
                full = ' '.join(parts).strip()
                if full:
                    return full.replace(',', ' •')
        return f"шир. {lat_fixed} • долг. {lng_fixed}"
    except Exception as e:
        logging.error(f"reverse_geocode error: {e}")
        return f"шир. {round(float(lat),5)} • долг. {round(float(lng),5)}"

def fill_report_record(reg_id, reg_data, gosb_name):
    if not supabase:
        return
    purpose = reg_data['purpose']
    existing = supabase.table('report').select('id').eq('registration_id', reg_id).execute()
    if existing.data:
        logging.info(f"Запись для registration_id={reg_id} уже существует")
        return
    employee_data = {}
    if reg_data.get('employee_id'):
        try:
            emp_res = supabase.table('employees').select('tab_number, kic_pi').eq('id', reg_data['employee_id']).execute()
            if emp_res.data:
                employee_data = emp_res.data[0]
        except Exception as e:
            logging.error(f"Ошибка получения сотрудника: {e}")
    row = {
        'registration_id': reg_id,
        'timestamp': reg_data['timestamp'],
        'fio': reg_data['fio'],
        'gosb_name': gosb_name,
        'tab_number': employee_data.get('tab_number'),
        'subdivision': employee_data.get('kic_pi'),
        'fire_training': '0,5',
        'radio_comm': '0,5',
        'drills': '0,25'
    }
    if purpose == 'Модуль 2':
        row['block_training'] = '8'
    elif purpose == 'Контраварийная подготовка':
        row['emergency'] = '8'
    elif purpose == 'Модуль 1':
        row['module1'] = '40'
    elif purpose == 'ЕПП':
        row['epp'] = '8'
    try:
        supabase.table('report').insert(row).execute()
    except Exception as e:
        logging.error(f"fill_report_record error: {e}")

def is_duplicate_registration(fio, purpose):
    if not supabase:
        return False
    try:
        now = datetime.now()
        year = now.year
        quarter = (now.month - 1) // 3 + 1
        start_date = datetime(year, (quarter-1)*3 + 1, 1)
        end_date = datetime(year, quarter*3 + 1, 1) - timedelta(days=1)
        res = supabase.table('registrations') \
            .select('id') \
            .eq('fio', fio) \
            .eq('purpose', purpose) \
            .gte('timestamp', start_date.isoformat()) \
            .lte('timestamp', end_date.isoformat()) \
            .execute()
        return len(res.data) > 0
    except Exception as e:
        logging.error(f"Ошибка проверки дубликатов: {e}")
        return False

def send_telegram_message(chat_id, text):
    if not TELEGRAM_BOT_TOKEN:
        logging.error("TELEGRAM_BOT_TOKEN не задан")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'}
    try:
        r = requests.post(url, json=payload, timeout=10)
        r.raise_for_status()
        logging.info(f"Сообщение отправлено в {chat_id}")
    except Exception as e:
        logging.error(f"Ошибка отправки в Telegram {chat_id}: {e}")

def format_report_message(registrations):
    if not registrations:
        return None
    kic_groups = {}
    for reg in registrations:
        kic = reg.get('subdivision') or 'Без КИЦ'
        kic_groups.setdefault(kic, []).append(reg['fio'])
    lines = []
    for kic, fios in sorted(kic_groups.items()):
        lines.append(f"🟢 КИЦ {kic}")
        lines.extend(f"👮 {fio}" for fio in fios)
        lines.append("")
    total = len(registrations)
    lines.append(f"📨 Итого: {total} {pluralize(total, 'человек', 'человека', 'человек')}")
    return "\n".join(lines).strip()

def pluralize(n, one, few, many):
    n = abs(n) % 100
    n1 = n % 10
    if 10 < n < 20: return many
    if 1 < n1 < 5: return few
    if n1 == 1: return one
    return many

def send_telegram_to_gosb(gosb, message):
    if gosb.get('chat_id'):
        send_telegram_message(gosb['chat_id'], message)
    if gosb.get('copy_chat_id'):
        send_telegram_message(gosb['copy_chat_id'], message)

def send_email(recipient, subject, body, cc=None):
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASSWORD:
        logging.error("SMTP не настроен")
        return False
    try:
        if SMTP_PORT == 587:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT)
        server.login(SMTP_USER, SMTP_PASSWORD)
        msg = MIMEMultipart()
        msg['From'] = SMTP_FROM
        msg['To'] = recipient if isinstance(recipient, str) else ', '.join(recipient)
        if cc:
            msg['Cc'] = cc if isinstance(cc, str) else ', '.join(cc)
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        all_recipients = [recipient] if isinstance(recipient, str) else recipient.copy()
        if cc:
            all_recipients.extend([cc] if isinstance(cc, str) else cc)
        server.send_message(msg, to_addrs=all_recipients)
        server.quit()
        logging.info(f"Email отправлен на {recipient}")
        return True
    except Exception as e:
        logging.error(f"SMTP error: {e}")
        return False

def create_excel_from_data(data_rows):
    if not OPENPYXL_AVAILABLE:
        raise Exception("openpyxl not installed")
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Дата и время", "Табельный №", "Ф.И.О.", "Подразделение",
        "Огневая ОЭТ, VR- тренажер/(УТС автомат)", "Блочное обучение",
        "Радиосвязь и мониторинг", "Учения по мониторингу и взаимодействию с СЦ ЦМИ",
        "Контраварийная подготовка", "Модуль 1", "ЕПП"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== МАРШРУТЫ ==========
@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/register/<slug>')
def register_form(slug):
    gosb = get_gosb_by_slug(slug)
    if not gosb:
        return "ГОСБ не найден", 404
    cities = get_cities_by_gosb(gosb['id'])
    return render_template('register.html', gosb=gosb, cities=cities)

@app.route('/api/register', methods=['POST'])
def api_register():
    if not supabase:
        return jsonify({'status': 'error', 'message': 'База данных не настроена'}), 500
    data = request.get_json()
    fio = data.get('fio')
    # Поддержка city_id (число) или cityId/city_name (строка)
    city_value = data.get('city_id') or data.get('cityId') or data.get('city_name')
    purpose = data.get('purpose')
    lat = data.get('latitude')
    lng = data.get('longitude')
    gosb_slug = data.get('gosb_slug')
    employee_id = data.get('employee_id') or data.get('employeeId')

    if not fio or not city_value or not purpose:
        logging.error(f"Missing fields: fio={fio}, city={city_value}, purpose={purpose}")
        return jsonify({'status': 'error', 'message': 'Не все поля заполнены'}), 400

    gosb = get_gosb_by_slug(gosb_slug)
    if not gosb:
        return jsonify({'status': 'error', 'message': 'ГОСБ не найден'}), 400

    # Определяем city_id
    if isinstance(city_value, str) and not city_value.isdigit():
        # Это название города
        city_res = supabase.table('cities').select('id').eq('name', city_value).eq('gosb_id', gosb['id']).execute()
        if not city_res.data:
            return jsonify({'status': 'error', 'message': f'Город "{city_value}" не найден в справочнике'}), 400
        city_id = city_res.data[0]['id']
    else:
        city_id = int(city_value)

    # Проверка дубликата
    if is_duplicate_registration(fio, purpose):
        return jsonify({
            'status': 'error',
            'message': f'❌ {fio} уже зарегистрирован на "{purpose}" в этом квартале.'
        }), 400

    address = reverse_geocode(lat, lng) if lat and lng else 'Адрес не определён'

    reg_data = {
        'fio': fio,
        'city_id': city_id,
        'purpose': purpose,
        'address': address,
        'timestamp': datetime.utcnow().isoformat(),
        'employee_id': int(employee_id) if employee_id and str(employee_id).isdigit() else None
    }
    try:
        result = supabase.table('registrations').insert(reg_data).execute()
        if not result.data:
            return jsonify({'status': 'error', 'message': 'Ошибка сохранения'}), 500
        reg_id = result.data[0]['id']
        fill_report_record(reg_id, reg_data, gosb['name'])
        return jsonify({
            'status': 'success',
            'message': 'Регистрация успешна',
            'address': address,
            'timestamp': datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        })
    except Exception as e:
        logging.error(f"Registration error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/gosb-list')
def get_gosb_list():
    if not supabase:
        return jsonify([])
    try:
        res = supabase.table('gosb').select('id, name').execute()
        return jsonify(res.data)
    except Exception as e:
        logging.error(f"gosb-list error: {e}")
        return jsonify([])

@app.route('/api/cities-by-gosb')
def get_cities_by_gosb_name():
    gosb_name = request.args.get('gosb_name')
    if not gosb_name or not supabase:
        return jsonify([])
    try:
        gosb_res = supabase.table('gosb').select('id').eq('name', gosb_name).execute()
        if not gosb_res.data:
            return jsonify([])
        gosb_id = gosb_res.data[0]['id']
        cities_res = supabase.table('cities').select('name').eq('gosb_id', gosb_id).execute()
        return jsonify([c['name'] for c in cities_res.data])
    except Exception as e:
        logging.error(f"cities-by-gosb error: {e}")
        return jsonify([])

@app.route('/api/report-data')
def get_report_data():
    if not supabase:
        return jsonify([])
    gosb = request.args.get('gosb')
    city = request.args.get('city')
    fio = request.args.get('fio')
    year = request.args.get('year')
    quarter = request.args.get('quarter')
    month = request.args.get('month')
    exact_date = request.args.get('exact_date')
    try:
        query = supabase.table('report').select('*')
        if gosb:
            query = query.eq('gosb_name', gosb)
        if city:
            query = query.ilike('subdivision', f'%{city}%')
        if fio:
            query = query.ilike('fio', f'%{fio}%')
        res = query.execute()
        data = res.data
        filtered = []
        for row in data:
            ts = row.get('timestamp')
            if not ts:
                continue
            try:
                utc_dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
            except:
                continue
            yekat_dt = utc_dt.replace(tzinfo=pytz.UTC).astimezone(YEKAT_TIMEZONE)
            formatted_date = yekat_dt.strftime('%d.%m.%Y %H:%M:%S')
            row['timestamp'] = formatted_date
            if exact_date and yekat_dt.date().isoformat() != exact_date:
                continue
            if year and str(yekat_dt.year) != year:
                continue
            if quarter:
                q = (yekat_dt.month - 1) // 3 + 1
                if str(q) != quarter:
                    continue
            if month and str(yekat_dt.month) != month:
                continue
            filtered.append(row)
        # дедупликация по registration_id
        unique = {}
        for row in filtered:
            rid = row.get('registration_id')
            if rid not in unique:
                unique[rid] = row
        filtered = list(unique.values())
        return jsonify(filtered)
    except Exception as e:
        logging.error(f"report-data error: {e}")
        return jsonify([])

@app.route('/api/employees/search')
def search_employees():
    if not supabase:
        return jsonify([])
    query = request.args.get('q', '').strip()
    limit = int(request.args.get('limit', 10))
    if not query:
        return jsonify([])
    try:
        if query.isdigit():
            res = supabase.table('employees').select('id, fio, tab_number, kic_pi').eq('tab_number', query).limit(limit).execute()
        else:
            res = supabase.table('employees').select('id, fio, tab_number, kic_pi').ilike('fio', f'%{query}%').limit(limit).execute()
        return jsonify(res.data)
    except Exception as e:
        logging.error(f"Ошибка поиска сотрудников: {e}")
        return jsonify([]), 500

@app.route('/api/statistics')
def get_statistics():
    if not supabase:
        return jsonify({"error": "База данных не инициализирована"}), 500
    gosb_name = request.args.get('gosb')
    city = request.args.get('city')
    year = request.args.get('year')
    quarter = request.args.get('quarter')
    month = request.args.get('month')
    exact_date = request.args.get('exact_date')
    emp_query = supabase.table('employees').select('fio, tab_number, kic_pi, gosb_name')
    if gosb_name:
        emp_query = emp_query.eq('gosb_name', gosb_name)
    if city:
        emp_query = emp_query.ilike('kic_pi', f'%{city}%')
    emp_res = emp_query.execute()
    employee_ids = set()
    for e in emp_res.data:
        if e.get('tab_number'):
            employee_ids.add(e['tab_number'])
        elif e.get('fio'):
            employee_ids.add(e['fio'].strip().lower())
    total_employees = len(employee_ids)
    report_query = supabase.table('report').select('fio, tab_number, subdivision, timestamp, registration_id')
    if gosb_name:
        report_query = report_query.eq('gosb_name', gosb_name)
    if city:
        report_query = report_query.ilike('subdivision', f'%{city}%')
    report_res = report_query.execute()
    unique_regs = {}
    for row in report_res.data:
        rid = row.get('registration_id')
        if rid and rid not in unique_regs:
            unique_regs[rid] = row
    unique_regs_list = list(unique_regs.values())
    filtered_regs = []
    for row in unique_regs_list:
        ts = row.get('timestamp')
        if not ts:
            continue
        try:
            utc_dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
        except:
            continue
        yekat_dt = utc_dt.replace(tzinfo=pytz.UTC).astimezone(YEKAT_TIMEZONE)
        if exact_date and yekat_dt.date().isoformat() != exact_date:
            continue
        if year and str(yekat_dt.year) != year:
            continue
        if quarter:
            q = (yekat_dt.month - 1) // 3 + 1
            if str(q) != quarter:
                continue
        if month and str(yekat_dt.month) != month:
            continue
        filtered_regs.append(row)
    registered_ids = set()
    for reg in filtered_regs:
        if reg.get('tab_number'):
            registered_ids.add(reg['tab_number'])
        elif reg.get('fio'):
            registered_ids.add(reg['fio'].strip().lower())
    registered_count = len(registered_ids)
    percentage = (registered_count / total_employees * 100) if total_employees > 0 else 0
    return jsonify({
        "total_employees": total_employees,
        "registered_unique": registered_count,
        "percentage": round(percentage, 1)
    })

# ========== НАПОМИНАНИЯ РУКОВОДИТЕЛЯМ КИЦ ==========
@app.route('/api/send-kic-reminders', methods=['POST'])
def send_kic_reminders():
    if not supabase:
        return jsonify({"error": "База данных не инициализирована"}), 500
    data = request.get_json() or {}
    days = data.get('days', 30)
    purpose = data.get('purpose')
    gosb = data.get('gosb')
    city = data.get('city')

    # 1. Сотрудники с фильтрацией
    emp_query = supabase.table('employees').select('id, fio, tab_number, kic_pi, gosb_name')
    if gosb:
        emp_query = emp_query.eq('gosb_name', gosb)
    if city:
        emp_query = emp_query.ilike('kic_pi', f'%{city}%')
    employees = emp_query.execute().data
    if not employees:
        return jsonify({"message": "Нет сотрудников для выбранных фильтров"}), 200

    # 2. Регистрации за период
    reg_query = supabase.table('registrations').select('employee_id, fio, purpose')
    if days > 0:
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        reg_query = reg_query.gte('timestamp', cutoff)
    if purpose:
        reg_query = reg_query.eq('purpose', purpose)
    reg_res = reg_query.execute()
    registered_ids = set()
    for reg in reg_res.data:
        if reg.get('employee_id'):
            registered_ids.add(reg['employee_id'])
        else:
            registered_ids.add(reg['fio'].strip().lower())

    # 3. Отбираем незарегистрированных
    missing_by_kic = defaultdict(list)
    for emp in employees:
        emp_id = emp.get('id')
        emp_tab = emp.get('tab_number')
        emp_fio = emp.get('fio', '').strip().lower()
        if emp_id and emp_id in registered_ids:
            continue
        if emp_tab and emp_tab in registered_ids:
            continue
        if emp_fio and emp_fio in registered_ids:
            continue
        kic = emp.get('kic_pi') or 'Без КИЦ'
        missing_by_kic[kic].append(emp)

    # 4. Получаем все города для поиска email
    cities_all = supabase.table('cities').select('name, manager_email, responsible_email').execute().data

    results = []
    for kic, emp_list in missing_by_kic.items():
        if city and city.lower() not in kic.lower():
            continue
        manager_email = None
        responsible_email = None
        for c in cities_all:
            if c['name'].lower() in kic.lower():
                manager_email = c.get('manager_email')
                responsible_email = c.get('responsible_email')
                break
        if not manager_email:
            results.append({"kic": kic, "error": "Нет email руководителя"})
            continue
        subject = f"Напоминание: {len(emp_list)} сотрудников не зарегистрировались на обучение ({kic})"
        body = f"Здравствуйте!\n\nСледующие сотрудники КИЦ «{kic}» не зарегистрировались на обучение за последние {days} дней:\n\n"
        for emp in emp_list:
            body += f"• {emp['fio']} (таб. {emp.get('tab_number', 'нет')})\n"
        body += "\nПожалуйста, организуйте их регистрацию в системе."
        if send_email(manager_email, subject, body, cc=responsible_email):
            results.append({"kic": kic, "sent": len(emp_list), "to": manager_email, "cc": responsible_email})
        else:
            results.append({"kic": kic, "error": "Ошибка отправки email"})

    if not results:
        return jsonify({"message": "Нет подходящих КИЦ для отправки"}), 200
    return jsonify({"status": "ok", "results": results})

# ========== ТЕСТОВЫЙ МАРШРУТ ДЛЯ SMTP ==========
@app.route('/api/test-email')
def test_email():
    result = send_email('vserosinkas@gmail.com', 'Тест SMTP', 'Если вы видите это письмо, SMTP работает правильно.')
    return jsonify({"status": "ok" if result else "error", "result": result})

# ========== ЭКСПОРТ В EXCEL ==========
@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "Excel export not available (openpyxl missing)"}), 500
    data = request.get_json()
    rows = data.get('data', [])
    if not rows:
        return jsonify({"error": "Нет данных для экспорта"}), 400
    try:
        excel_file = create_excel_from_data(rows)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logging.error(f"Export error: {e}")
        return jsonify({"error": str(e)}), 500

# ========== ОТПРАВКА В TELEGRAM ==========
@app.route('/api/send-daily-reports', methods=['GET'])
def send_daily_reports():
    if not supabase:
        return jsonify({"error": "База данных не инициализирована"}), 500
    gosb_res = supabase.table('gosb').select('id, name, slug, chat_id, copy_chat_id').not_.is_('chat_id', 'null').execute()
    if not gosb_res.data:
        return jsonify({"message": "Нет получателей"}), 200
    yesterday = date.today() - timedelta(days=1)
    yesterday_str = yesterday.strftime('%Y-%m-%d')
    report_res = supabase.table('report').select('*').gte('timestamp', yesterday_str).lte('timestamp', yesterday_str + ' 23:59:59').execute()
    registrations = report_res.data
    by_gosb = {}
    for reg in registrations:
        by_gosb.setdefault(reg.get('gosb_name'), []).append(reg)
    sent_count = 0
    for gosb in gosb_res.data:
        regs = by_gosb.get(gosb['name'], [])
        if not regs:
            continue
        message = format_report_message(regs)
        if message:
            full_message = f"🏢 {gosb['name']} — регистрация на {yesterday.strftime('%d.%m.%Y')}\n\n{message}"
            send_telegram_to_gosb(gosb, full_message)
            sent_count += 1
    return jsonify({"status": "ok", "sent": sent_count}), 200

@app.route('/api/send-today-reports', methods=['POST'])
def send_today_reports():
    if not supabase:
        return jsonify({"error": "База данных не инициализирована"}), 500
    gosb_res = supabase.table('gosb').select('id, name, slug, chat_id, copy_chat_id').not_.is_('chat_id', 'null').execute()
    if not gosb_res.data:
        return jsonify({"message": "Нет получателей"}), 200
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    report_res = supabase.table('report').select('*').gte('timestamp', today_str).lte('timestamp', today_str + ' 23:59:59').execute()
    registrations = report_res.data
    by_gosb = {}
    for reg in registrations:
        by_gosb.setdefault(reg.get('gosb_name'), []).append(reg)
    sent_count = 0
    for gosb in gosb_res.data:
        regs = by_gosb.get(gosb['name'], [])
        if not regs:
            continue
        message = format_report_message(regs)
        if message:
            full_message = f"🏢 {gosb['name']} — регистрация на {today.strftime('%d.%m.%Y')}\n\n{message}"
            send_telegram_to_gosb(gosb, full_message)
            sent_count += 1
    return jsonify({"status": "ok", "sent": sent_count}), 200

@app.route('/api/cron-remind', methods=['GET'])
def cron_remind():
    with app.test_request_context(json={'days': 30}):
        return send_kic_reminders()

@app.route('/api/debug-supabase')
def debug_supabase():
    if not supabase:
        return jsonify({"error": "Supabase client not initialized"}), 500
    try:
        res = supabase.table('gosb').select('*').limit(1).execute()
        return jsonify({"status": "ok", "data": res.data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
